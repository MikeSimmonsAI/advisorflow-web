"""READ-ONLY probe: how does the Opportunities export relate to the contact master?"""
import sys
from collections import Counter
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from openpyxl import load_workbook

MASTER = r"C:\Users\simmo\Downloads\All Active Sales Leads (Excludes Non-Viable) 8-31-2026 7-50-01 PM.xlsx"
OPPS = r"C:\Users\simmo\Downloads\All Opportunities 8-31-2026 8-20-32 PM.xlsx"


def load(path, cols):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [("" if h is None else str(h).strip()) for h in next(it)]
    idx = {h: i for i, h in enumerate(header)}
    want = {c: idx[c] for c in cols if c in idx}
    rows = []
    for r in it:
        if r is None:
            continue
        rows.append({c: r[i] for c, i in want.items()})
    wb.close()
    return header, rows


print("loading master contact keys...")
_, m = load(MASTER, ["(Do Not Modify) Contact", "Full Name", "Status Reason",
                     "Sale Made?", "Last Sold Date"])
master_ids = {str(r["(Do Not Modify) Contact"]).strip() for r in m
              if r.get("(Do Not Modify) Contact")}
print("master rows:", len(m), " distinct contact ids:", len(master_ids))

OPP_COLS = ["(Do Not Modify) Opportunity", "Opportunity ID", "Lead ID", "LeadID",
            "Contact", "Status", "Status Reason", "Contract #", "Contract Date",
            "Contract Total", "Actual Revenue", "Actual Close Date", "Date Signed",
            "Contract Cancelled", "Contract Close Status", "Need", "Contract Need",
            "Sales Advisor", "Location", "Purchaser", "Contract Type"]
print("loading opportunities...")
oh, o = load(OPPS, OPP_COLS)
print("opportunity rows:", len(o))


def fill(col):
    n = sum(1 for r in o if r.get(col) is not None and str(r.get(col)).strip())
    print(f"  {col:<32} {n:>7}  ({n/len(o)*100:5.1f}%)")


print("\nFILL RATES")
for c in OPP_COLS:
    if c in oh:
        fill(c)
    else:
        print(f"  {c:<32} (absent)")

print("\nJOIN TESTS against master '(Do Not Modify) Contact'")
for key in ["Contact", "Lead ID", "LeadID", "Opportunity ID",
            "(Do Not Modify) Opportunity"]:
    if key not in oh:
        continue
    vals = [str(r[key]).strip() for r in o
            if r.get(key) is not None and str(r.get(key)).strip()]
    distinct = set(vals)
    hit = len(distinct & master_ids)
    sample = next(iter(distinct), "")
    print(f"  {key:<30} nonblank={len(vals):>7} distinct={len(distinct):>7} "
          f"in master={hit:>7} ({hit/max(len(distinct),1)*100:5.1f}%) "
          f"sample_len={len(sample)}")

# cardinality on the best key
best = None
bestn = -1
for key in ["Contact", "Lead ID", "LeadID"]:
    if key not in oh:
        continue
    distinct = {str(r[key]).strip() for r in o if r.get(key) and str(r[key]).strip()}
    hit = len(distinct & master_ids)
    if hit > bestn:
        bestn, best = hit, key
print("\nbest join key:", best, "->", bestn, "master contacts matched")
if best:
    c = Counter(str(r[best]).strip() for r in o if r.get(best) and str(r[best]).strip())
    dist = Counter(c.values())
    print("  opportunities per key:")
    for k in sorted(dist)[:10]:
        print(f"    {k} opportunity(ies): {dist[k]} keys")
    print("  max per key:", max(c.values()) if c else 0)

print("\nVALUE DISTRIBUTIONS")
for col in ["Status", "Status Reason", "Contract Close Status", "Contract Cancelled",
            "Need", "Contract Need", "Contract Type"]:
    if col not in oh:
        continue
    cc = Counter(("" if r.get(col) is None else str(r[col]).strip())[:38] for r in o)
    print(f"  === {col}")
    for v, n in cc.most_common(10):
        print(f"     {n:>7}  {v!r}")

print("\nMONEY / DATE SANITY")
for col in ["Contract Total", "Actual Revenue"]:
    if col not in oh:
        continue
    nums = []
    for r in o:
        v = r.get(col)
        if v is None:
            continue
        try:
            nums.append(float(str(v).replace("$", "").replace(",", "")))
        except ValueError:
            pass
    if nums:
        nums.sort()
        nz = [x for x in nums if x]
        print(f"  {col}: n={len(nums)} nonzero={len(nz)} "
              f"min={nums[0]:.0f} median={nums[len(nums)//2]:.0f} max={nums[-1]:.0f}")
print("DONE")
