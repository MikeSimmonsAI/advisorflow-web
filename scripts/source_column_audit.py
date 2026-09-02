"""
SOURCE COLUMN AUDIT - every column of a source file, classified.

    python scripts/source_column_audit.py <file> [--sheet N] [--csv out.csv]

For each column reports:

    PURPOSE       COMPLIANCE | HISTORICAL_ACTIVITY | IDENTITY | PROVENANCE |
                  COMMERCIAL | OTHER
    STAGED        where it lands when staged as a historical source record
    OPERATIONAL   where it lands when imported as leads
                  MAPPED | CUSTOM_FIELDS_ONLY

Read-only: opens the file, reads its header row, closes it.
"""

import argparse
import csv
import os
import sys
from collections import Counter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from app.services.import_service import _build_column_lookup, HEADER_MAP  # noqa: E402
from app.services.source_ingest import (classify_columns, DESTINATIONS,     # noqa: E402
                                        PURPOSES, MAPPED, CUSTOM_FIELDS_ONLY)
from app.services import permission_values as pv                            # noqa: E402


def header_of(path: str, sheet: int = 0) -> list[str]:
    low = path.lower()
    if low.endswith(".csv") or low.endswith(".tsv"):
        with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
            return next(csv.reader(f, delimiter="\t" if low.endswith(".tsv") else ","))
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[sheet]]
    header = [("" if h is None else str(h).strip())
              for h in next(ws.iter_rows(values_only=True))]
    wb.close()
    return header


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--sheet", type=int, default=0)
    ap.add_argument("--csv", default="")
    a = ap.parse_args()

    header = header_of(a.file, a.sheet)
    lookup = _build_column_lookup(header)
    rows = classify_columns(header, operational_lookup=lookup)

    print(f"FILE: {os.path.basename(a.file)}")
    print(f"COLUMNS: {len(header)}\n")
    print(f"{'COLUMN':<42} {'PURPOSE':<20} {'STAGED':<19} OPERATIONAL")
    print("-" * 104)
    for r in rows:
        print(f"{r['column']:<42} {r['purpose']:<20} "
              f"{r['staged_destination']:<19} {r['operational_destination']}")

    by_purpose = Counter(r["purpose"] for r in rows)
    by_staged = Counter(r["staged_destination"] for r in rows)
    by_op = Counter(r["operational_destination"] for r in rows)

    print("\nBY PURPOSE")
    for k in PURPOSES:
        print(f"  {k:<22} {by_purpose.get(k, 0)}")
    print("\nSTAGED DESTINATION")
    for k in DESTINATIONS:
        print(f"  {k:<22} {by_staged.get(k, 0)}")
    print("\nOPERATIONAL DESTINATION")
    for k in DESTINATIONS:
        print(f"  {k:<22} {by_op.get(k, 0)}")

    print("\nCANONICAL FIELDS THE OPERATIONAL IMPORTER RESOLVED")
    for k in HEADER_MAP:
        print(f"  {k:<22} {lookup.get(k, '-- NOT MAPPED --')}")

    print("\nCOMPLIANCE COLUMNS PRESENT")
    comp = [r for r in rows if r["purpose"] == "COMPLIANCE"]
    if not comp:
        print("  (none) - this file states no channel permission at all")
    for r in comp:
        low = r["column"].lower()
        which = [p for p in pv.PERMISSIONS
                 if low in {c for c, _ in pv.COLUMN_TABLE[p]}]
        pol = [pol for p in which for c, pol in pv.COLUMN_TABLE[p] if c == low]
        print(f"  {r['column']:<42} -> {', '.join(which)}  (bare-boolean polarity: "
              f"{pol[0] if pol else '-'})")
        if r["operational_destination"] != MAPPED:
            print("     *** COMPLIANCE COLUMN NOT MAPPED BY THE IMPORTER ***")

    print("\nHISTORICAL ACTIVITY COLUMNS PRESENT")
    hist = [r for r in rows if r["purpose"] == "HISTORICAL_ACTIVITY"]
    for r in hist:
        print(f"  {r['column']:<42} {r['operational_destination']}")

    # WHAT COUNTS AS A FAILURE, AND WHAT DOES NOT.
    #
    # A COMPLIANCE column that the importer parks is a failure, always: an
    # opt-out inside a JSON blob is an opt-out no send path will ever read.
    #
    # A SECONDARY activity column that the importer parks is NOT a failure. A
    # Lead carries one `last_contact_date`; a file with six date columns cannot
    # map all six, and the staged source record keeps them anyway. What WOULD
    # be a failure is a file that offers an activity date and an importer that
    # maps none of them - which is the defect that started this work.
    parked_compliance = [r for r in rows
                         if r["purpose"] == "COMPLIANCE"
                         and r["operational_destination"] == CUSTOM_FIELDS_ONLY]
    activity_columns = [r for r in rows if r["purpose"] == "HISTORICAL_ACTIVITY"]
    activity_mapped = "last_contact_date" in lookup
    secondary = [r for r in activity_columns
                 if r["operational_destination"] == CUSTOM_FIELDS_ONLY]

    print(f"\nCOMPLIANCE COLUMNS THE IMPORTER WOULD PARK: {len(parked_compliance)}"
          + ("  <-- FAILURE" if parked_compliance else "  (none)"))
    for r in parked_compliance:
        print(f"  {r['column']}")

    print(f"\nHISTORICAL CONTACT DATE MAPPED: "
          f"{lookup.get('last_contact_date', 'NO')}")
    if activity_columns and not activity_mapped:
        print("  <-- FAILURE: this file offers activity columns and none is mapped")
    print(f"SECONDARY ACTIVITY COLUMNS (staged, not on the lead row): {len(secondary)}")
    for r in secondary:
        print(f"  {r['column']}")

    lost = parked_compliance or (activity_columns and not activity_mapped)

    if a.csv:
        with open(a.csv, "w", encoding="utf-8", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=["column", "purpose",
                                               "staged_destination",
                                               "operational_destination"])
            w.writeheader()
            w.writerows(rows)
        print(f"\nwritten: {a.csv}")

    # A compliance or activity column that the importer parks is a FAILURE of
    # this audit, not a note in it.
    return 1 if lost else 0


if __name__ == "__main__":
    raise SystemExit(main())
