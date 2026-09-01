"""
WHAT WILL THE IMPORTER ACTUALLY KEEP FROM THIS FILE?

    python scripts/import_mapping_report.py <file> [<file> ...]

`import_service._build_column_lookup` matches headers EXACTLY against a fixed
alias list. A header the list does not name is not "matched anyway" - it is
parked in `custom_fields`, where nothing downstream reads it. That is invisible
at upload time: the import reports success, the rows appear, and the evidence
is simply not on them.

This report makes it visible BEFORE an import, and shows what the reconciliation
adapter would additionally read from the same file. Read-only; opens the file,
reads its header row, and touches nothing else.
"""

import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from app.services.import_service import _build_column_lookup, HEADER_MAP   # noqa: E402
from app.services.source_adapters import (ALIASES, DATE_FIELDS,            # noqa: E402
                                          EMAIL_FIELDS, PHONE_FIELDS,
                                          PERMISSION_FIELDS)


def header_of(path: str, sheet: int = 0) -> list[str]:
    low = path.lower()
    if low.endswith(".csv") or low.endswith(".tsv"):
        import csv
        with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
            return next(csv.reader(f, delimiter="\t" if low.endswith(".tsv") else ","))
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[sheet]]
    header = [("" if h is None else str(h).strip())
              for h in next(ws.iter_rows(values_only=True))]
    wb.close()
    return header


def adapter_columns() -> set[str]:
    known: set[str] = set()
    for names in ALIASES.values():
        known |= set(names)
    known |= set(DATE_FIELDS) | set(EMAIL_FIELDS) | set(PHONE_FIELDS)
    for cols in PERMISSION_FIELDS.values():
        known |= {c for c, _ in cols}
    return known


def report(path: str, sheet: int) -> None:
    header = header_of(path, sheet)
    lookup = _build_column_lookup(header)
    mapped = set(lookup.values())

    print("=" * 74)
    print(os.path.basename(path))
    print(f"  columns: {len(header)}")

    missing = [k for k in HEADER_MAP if k not in lookup]
    print(f"  importer maps {len(lookup)} of {len(HEADER_MAP)} known fields")
    for k in HEADER_MAP:
        print(f"    {k:<20} {lookup.get(k, '-- NOT MAPPED --')}")
    if missing:
        print(f"  NOT MAPPED: {', '.join(missing)}")

    parked = [h for h in header if h and h not in mapped]
    print(f"  parked into custom_fields: {len(parked)}")
    for h in parked:
        print(f"    {h}")

    low = {h.lower() for h in header if h}
    read_by_adapter = low & adapter_columns()
    extra = sorted(read_by_adapter - {c.lower() for c in mapped})
    print(f"  reconciliation adapter reads {len(read_by_adapter)} of {len(header)}")
    print(f"  columns the adapter reads that the importer parks: {len(extra)}")
    for h in extra:
        print(f"    {h}")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("files", nargs="+")
    ap.add_argument("--sheet", type=int, default=0)
    a = ap.parse_args()
    for p in a.files:
        try:
            report(p, a.sheet)
        except Exception as exc:
            print(f"{p}: {type(exc).__name__}: {exc}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
