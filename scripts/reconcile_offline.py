"""
READ-ONLY offline reconciliation runner.

    python scripts/reconcile_offline.py --target <file> --source <file> [--out <csv>]

Compares one operational population against one historical source using the
platform engine in app/services/source_reconciliation.py. It opens both files
read-only, writes NOTHING back to either, holds no database handle, and has no
send path of any kind.

It takes both populations as ARGUMENTS. No customer, advisor, file name or
population size is written into this script.
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
from collections import Counter
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from app.services.source_adapters import rows_to_records          # noqa: E402
from app.services import source_reconciliation as sr              # noqa: E402


def read_table(path: str, sheet: int = 0) -> list[dict]:
    low = path.lower()
    if low.endswith(".csv") or low.endswith(".tsv"):
        delim = "\t" if low.endswith(".tsv") else ","
        with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
            return [dict(r) for r in csv.DictReader(f, delimiter=delim)]
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[sheet]]
    it = ws.iter_rows(values_only=True)
    header = [("" if h is None else str(h).strip()) for h in next(it)]
    rows = []
    for r in it:
        if r is None:
            continue
        d = {header[i]: r[i] for i in range(min(len(header), len(r)))}
        if any(v is not None and str(v).strip() != "" for v in d.values()):
            rows.append(d)
    wb.close()
    return rows


def mask_email(e: str) -> str:
    if not e or "@" not in e:
        return ""
    user, dom = e.split("@", 1)
    keep = user[:1]
    return f"{keep}{'*' * max(len(user) - 1, 1)}@{dom}"


def mask_phone(p: str) -> str:
    d = "".join(ch for ch in (p or "") if ch.isdigit())
    return f"***-***-{d[-4:]}" if len(d) >= 4 else ""


# Ranking is EVIDENCE ORDER, not a score. Each key is a fact on the record.
DISPOSITION_RANK = {
    "appointment set": 5, "presentation made": 4, "contacted": 3,
    "attempting contact": 2, "new": 0,
}


def rank_key(f: dict):
    t, s = f["target"], f["source"]
    reason = ((s.status_reason if s else "") or t.status_reason or "").strip().lower()
    disp = DISPOSITION_RANK.get(reason, 1 if reason and reason != "new" else 0)
    action = 1 if ((s.last_action if s else "") or t.last_action).strip() else 0
    date = (s.last_contact_date if s else None) or t.last_contact_date
    recency = date.timestamp() if isinstance(date, datetime) else 0.0
    reach = int(bool(f["_email"])) + int(bool(f["_phone"]))
    return (-disp, -action, -recency, -reach)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--target", required=True, help="operational population file")
    ap.add_argument("--source", required=True, help="historical master file")
    ap.add_argument("--target-sheet", type=int, default=0)
    ap.add_argument("--source-sheet", type=int, default=0)
    ap.add_argument("--top", type=int, default=25)
    ap.add_argument("--channel", default="",
                    choices=("", "email", "bulk_email", "sms", "voice"),
                    help="rank only records this channel is not restricted for")
    ap.add_argument("--out", default="", help="optional CSV of the full findings")
    a = ap.parse_args()

    print("READ-ONLY RECONCILIATION")
    print("target:", os.path.basename(a.target))
    print("source:", os.path.basename(a.source))

    t_rows = read_table(a.target, a.target_sheet)
    s_rows = read_table(a.source, a.source_sheet)
    print(f"target rows: {len(t_rows):,}   source rows: {len(s_rows):,}")

    targets = rows_to_records(t_rows)
    for i, r in enumerate(targets):
        if not r.key:
            r.key = f"row-{i + 1}"
    sources = rows_to_records(s_rows)

    index = sr.SourceIndex(sources)
    print(f"index: source_key={len(index.by_source_key):,} "
          f"email={len(index.by_email):,} phone={len(index.by_phone):,} "
          f"name={len(index.by_name):,}")

    findings = sr.reconcile(targets, index)

    # resolved contactability, used for reporting only
    for f in findings:
        t = f["target"]
        fills = {x["field"]: x["proposed"] for x in f["enrichment"]["fills"]}
        f["_email"] = t.email or sr.normalize_email(str(fills.get("email", "")))
        f["_phone"] = t.phone or fills.get("phone", "")

    s = sr.summarize(findings)

    print("\n=== RECONCILIATION REPORT ===")
    print(f"TOTAL CURRENT RECORDS                 {s['total']}")
    for k in sr.MATCH_STATUSES:
        print(f"{k:<37} {s['match_status'][k]}")
    print("\nMATCH RULES THAT FIRED")
    for k, v in sorted(s["match_rules"].items(), key=lambda x: -x[1]):
        conf = next((c for c, n, _ in sr.MATCH_RULES if c == k), "")
        del conf
        print(f"  {k:<35} {v}")

    print("\nFIELDS AVAILABLE FOR ENRICHMENT (blank-fill candidates)")
    if not s["enrichment_fills"]:
        print("  (none)")
    for k, v in sorted(s["enrichment_fills"].items(), key=lambda x: -x[1]):
        print(f"  {k:<35} {v}")

    print("\nCONFLICTS (current populated and historical disagrees)")
    if not s["conflicts"]:
        print("  (none)")
    for k, v in sorted(s["conflicts"].items(), key=lambda x: -x[1]):
        print(f"  {k:<35} {v}")

    print(f"\nCOMPLIANCE RESTRICTIONS DISCOVERED     {s['compliance_restrictions_discovered']}")
    ch = Counter()
    for f in findings:
        for item in f["compliance"]["findings"]:
            ch[(item["channel"], item["reason_code"])] += 1
    for (c, code), n in ch.most_common():
        print(f"  {c:<11} {code:<50} {n}")

    print("\nRESOLVED CHANNEL PERMISSION (after more-restrictive-wins)")
    for c in sr.CHANNELS:
        st = Counter(f["compliance"]["channels"].get(c) for f in findings)
        print(f"  {c:<11} allow={st.get(True, 0):<5} deny={st.get(False, 0):<5} "
              f"not stated={st.get(None, 0)}")

    print("\nVIABILITY")
    for k in sr.VIABILITY_CLASSES:
        print(f"  {k:<35} {s['viability'][k]}")

    print("\nVIABILITY REASONS")
    rc = Counter()
    for f in findings:
        for r in f["viability_reasons"]:
            rc[r] += 1
    for k, v in rc.most_common(30):
        print(f"  {k:<45} {v}")

    workable = [f for f in findings
                if f["viability"] in (sr.VIABLE_READY, sr.VIABLE_LOWER_PRIORITY)]
    blocked_for_channel = 0
    if a.channel:
        # A person who is viable is not thereby reachable ON THIS CHANNEL.
        # Ranking them into a channel's work list would put somebody who opted
        # out of that channel at the top of it.
        before = len(workable)
        workable = [f for f in workable
                    if f["compliance"]["channels"].get(a.channel) is not False]
        blocked_for_channel = before - len(workable)
        print(f"\nCHANNEL FILTER '{a.channel}': {blocked_for_channel} viable "
              f"record(s) removed because that channel is restricted for them")
    workable.sort(key=rank_key)
    print(f"\n=== TOP {a.top} TO WORK  (of {len(workable)} workable) ===")
    for i, f in enumerate(workable[:a.top], 1):
        t, src = f["target"], f["source"]
        name = f"{t.first_name} {t.last_name}".strip() or "(no name)"
        date = (src.last_contact_date if src else None) or t.last_contact_date
        ds = date.strftime("%Y-%m-%d") if isinstance(date, datetime) else "—"
        reason = ((src.status_reason if src else "") or t.status_reason or "—")
        action = ((src.last_action if src else "") or t.last_action or "—")
        print(f"{i:>3}. {name:<28} {mask_email(f['_email']):<26} "
              f"{mask_phone(f['_phone']):<14} {f['viability']}")
        print(f"     match {f['match_rule']} ({f['match_confidence']})  "
              f"last activity {ds}  status {reason!r}  action {action!r}")
        print(f"     evidence: {', '.join(f['viability_reasons']) or '—'}")
        if f["enrichment"]["fills"]:
            print(f"     enrich:   {', '.join(x['field'] for x in f['enrichment']['fills'])}")

    if a.out:
        with open(a.out, "w", encoding="utf-8", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["key", "first_name", "last_name", "email", "phone",
                        "match_status", "match_rule", "match_confidence",
                        "source_key", "alternates", "viability",
                        "viability_reasons", "enrich_fields", "conflict_fields",
                        "compliance_findings", "hist_status_reason",
                        "hist_last_action", "hist_last_activity",
                        "resolved_email", "resolved_bulk_email", "resolved_sms",
                        "resolved_voice", "src_allow_emails", "src_bulk",
                        "src_text", "src_calls"])
            for f in findings:
                t, src = f["target"], f["source"]
                w.writerow([
                    f["key"], t.first_name, t.last_name, f["_email"], f["_phone"],
                    f["match_status"], f["match_rule"], f["match_confidence"],
                    f["source_key"] or "", f["alternate_count"], f["viability"],
                    "; ".join(f["viability_reasons"]),
                    "; ".join(x["field"] for x in f["enrichment"]["fills"]),
                    "; ".join(x["field"] for x in f["enrichment"]["conflicts"]),
                    "; ".join(f"{i['channel']}:{i['reason_code']}"
                              for i in f["compliance"]["findings"]),
                    (src.status_reason if src else ""),
                    (src.last_action if src else ""),
                    ((src.last_contact_date if src else None) or "").__str__(),
                    f["compliance"]["channels"].get("email"),
                    f["compliance"]["channels"].get("bulk_email"),
                    f["compliance"]["channels"].get("sms"),
                    f["compliance"]["channels"].get("voice"),
                    (src.raw.get("allow emails?") if src else ""),
                    (src.raw.get("do not allow bulk emails") if src else ""),
                    (src.raw.get("allow text message?") if src else ""),
                    (src.raw.get("allow phone calls?") if src else ""),
                ])
        print(f"\nfull findings written to {a.out}")

    print("\nNOTHING WAS MODIFIED. NO EMAIL, SMS, VOICE OR CADENCE WAS TOUCHED.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
