"""
READ-ONLY BUSINESS RECONCILIATION.

    python scripts/reconcile_business.py --target <file> --source <file>
                                         [--opportunities <file>] [--out <csv>]

Reconciles one confirmed operational population against one historical master
using the platform engine in app/services/source_reconciliation.py, then adds
the BUSINESS layer on top: canonical permission resolution, recommended
channel, and exactly one work classification per record.

It opens every file read-only, writes NOTHING back to any of them, holds no
database handle, imports no send path, and takes both populations as
arguments. No customer, advisor, file name or population size is written into
this script.

THE TWO FACTS THIS PASS IS BUILT AROUND
---------------------------------------
1. The operational rows carry NO canonical last-activity date. The original
   import parked that column. So the current record saying "never contacted"
   is an artefact of the import, not a fact about the family, and this pass
   takes activity evidence from the historical master instead.

2. The operational rows carry NO permission state at all, because the original
   import did not map the permission columns. Absent is not permitted. Every
   channel therefore starts UNKNOWN and is resolved from the master under
   more-restrictive-wins; UNKNOWN never becomes consent.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from app.services.source_adapters import rows_to_records            # noqa: E402
from app.services import source_reconciliation as sr                # noqa: E402
from app.services import permission_values as pv                    # noqa: E402


# ---------------------------------------------------------------------------
# Work classifications - exactly one per record, decided in this order.
# Anything that forbids or invalidates outreach is decided before anything
# that ranks it, so nothing is promoted past a restriction by scoring well.
# ---------------------------------------------------------------------------

DO_NOT_CONTACT = "DO_NOT_CONTACT"
DUPLICATE_REVIEW = "DUPLICATE_REVIEW"
ALREADY_RESOLVED = "ALREADY_RESOLVED"
BAD_CONTACT_DATA = "BAD_CONTACT_DATA"
NO_CONFIDENT_MASTER_MATCH = "NO_CONFIDENT_MASTER_MATCH"
REVIEW_REQUIRED = "REVIEW_REQUIRED"
WORK_NOW_EMAIL = "WORK_NOW_EMAIL"
WORK_NOW_VOICE = "WORK_NOW_VOICE"
WORK_LATER = "WORK_LATER"

WORK_CLASSES = (
    WORK_NOW_EMAIL, WORK_NOW_VOICE, WORK_LATER, REVIEW_REQUIRED,
    DO_NOT_CONTACT, DUPLICATE_REVIEW, ALREADY_RESOLVED, BAD_CONTACT_DATA,
    NO_CONFIDENT_MASTER_MATCH,
)

# Evidence order for ranking. These are facts on the record, not a score.
DISPOSITION_RANK = {
    "appointment set": 5,
    "presentation made": 4,
    "contacted": 3,
    "attempting contact": 2,
    "new": 0,
}


def read_table(path: str, sheet: int = 0) -> list[dict]:
    low = path.lower()
    if low.endswith(".csv") or low.endswith(".tsv"):
        delim = "\t" if low.endswith(".tsv") else ","
        with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
            return [dict(r) for r in csv.DictReader(f, delimiter=delim)]
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[sheet]]
    it = ws.iter_rows(values_only=True)
    header = [("" if h is None else str(h).strip()) for h in next(it)]
    rows = []
    for r in it:
        if r is None:
            continue
        d = {header[i]: r[i] for i in range(min(len(header), len(r)))}
        if any(v is not None and str(v).strip() != "" for v in d.values()):
            rows.append(d)
    wb.close()
    return rows


def mask_email(e: str) -> str:
    if not e or "@" not in e:
        return ""
    user, dom = e.split("@", 1)
    return f"{user[:1]}{'*' * max(len(user) - 1, 1)}@{dom}"


def mask_phone(p: str) -> str:
    d = "".join(ch for ch in (p or "") if ch.isdigit())
    return f"***-***-{d[-4:]}" if len(d) >= 4 else ""


def norm_name(s: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z ]", " ", (s or "").lower())).strip()


def ds(d) -> str:
    return d.strftime("%Y-%m-%d") if isinstance(d, datetime) else ""


def perm_word(v) -> str:
    return "ALLOWED" if v is True else "DENIED" if v is False else "UNKNOWN"


def classify_work(f: dict, ch: dict, has_email: bool, has_phone: bool,
                  opp_won: bool) -> tuple[str, list[str]]:
    """Exactly one primary classification, decided in a fixed order."""
    via = f["viability"]
    reasons = list(f["viability_reasons"])

    # 1. Forbidden - never overridden by anything below.
    if via == sr.DO_NOT_CONTACT:
        return DO_NOT_CONTACT, reasons

    # 2. Ambiguous identity. Two distinct historical people behind one
    #    operational row is a merge risk, not a work item.
    if f["match_status"] == sr.MULTIPLE_MATCHES or via == sr.DUPLICATE:
        return DUPLICATE_REVIEW, reasons or ["multiple_distinct_source_records"]

    # 3. Already dealt with - on the MASTER's authority only.
    #
    # The opportunity export identifies its contact by display NAME and
    # nothing else. On a population this size a common name matches many
    # unrelated opportunities (one record here name-matched twenty-one), and
    # the master states Sale Made = No for every one of them. Letting a name
    # collision overrule the master would retire live families as "already
    # sold" and quietly delete them from the work list. So the opportunity
    # join is reported on its own line and is never decisive by itself; it
    # only corroborates a sale the master already states.
    if via == sr.ALREADY_RESOLVED:
        return ALREADY_RESOLVED, reasons + (
            ["corroborated_by_opportunity_name_match"] if opp_won else [])

    # 4. Unreachable.
    if via == sr.BAD_CONTACT_DATA or not (has_email or has_phone):
        return BAD_CONTACT_DATA, reasons or ["no_usable_email_or_phone"]

    # 5. Identity not established against the master. Permissions cannot be
    #    resolved for a record with no historical counterpart, and an
    #    unresolved permission is not a permission.
    if f["match_status"] == sr.NO_MATCH:
        return NO_CONFIDENT_MASTER_MATCH, reasons or ["no_candidate_in_source"]

    # 6. Matched, but not strongly enough to act on without a human.
    if via == sr.REVIEW_REQUIRED or f["match_status"] == sr.MATCHED_REVIEW:
        return REVIEW_REQUIRED, reasons

    # 7. Workable. Channel is chosen from RESOLVED permission only.
    #    UNKNOWN is not a channel you may use.
    email_ok = ch.get("email") is True and has_email
    voice_ok = ch.get("voice") is True and has_phone
    sms_ok = ch.get("sms") is True and has_phone
    # SMS-only reachability has no WORK_NOW class in this vocabulary, so it
    # lands in review rather than being silently dropped - and the reason says
    # which channel IS permitted, because "no permitted channel" would be a
    # false statement about somebody who may lawfully be texted.
    tail = (["only_permitted_channel_is_sms"] if sms_ok
            else ["no_permitted_channel"])
    if via == sr.VIABLE_READY:
        if email_ok:
            return WORK_NOW_EMAIL, reasons
        if voice_ok:
            return WORK_NOW_VOICE, reasons
        return REVIEW_REQUIRED, reasons + tail
    if email_ok or voice_ok:
        return WORK_LATER, reasons
    return REVIEW_REQUIRED, reasons + tail


def recommended_channel(ch: dict, has_email: bool, has_phone: bool) -> str:
    """Only a channel the resolved permission actually allows."""
    if ch.get("email") is True and has_email:
        return "email"
    if ch.get("voice") is True and has_phone:
        return "voice"
    if ch.get("sms") is True and has_phone:
        return "sms"
    return "none_permitted"


# Channel eligibility is reported INDEPENDENTLY of the work classification.
# A record can be held out of the email work list and still be lawfully
# reachable by text; collapsing the two loses that, and a lead that is
# "REVIEW_REQUIRED because email is denied" is not a lead with no channels.
EMAIL_ELIGIBLE = "EMAIL_ELIGIBLE"
SMS_ELIGIBLE = "SMS_ELIGIBLE"
VOICE_ELIGIBLE = "VOICE_ELIGIBLE"
CHANNEL_REVIEW = "CHANNEL_REVIEW"
CHANNEL_DNC = "DO_NOT_CONTACT"


def channel_eligibility(ch: dict, has_email: bool, has_phone: bool,
                        suppressed: bool, dnc: bool) -> list[str]:
    """Which channels this person may lawfully be reached on, per channel.

    DO_NOT_CONTACT here means every contactable channel is denied or the
    record is suppressed - it is NOT returned merely because one channel is
    denied. CHANNEL_REVIEW means nothing is affirmatively permitted and
    nothing is affirmatively denied either: unknown, which is not consent and
    is also not a prohibition.
    """
    if suppressed or dnc:
        return [CHANNEL_DNC]
    out = []
    if ch.get("email") is True and has_email:
        out.append(EMAIL_ELIGIBLE)
    if ch.get("sms") is True and has_phone:
        out.append(SMS_ELIGIBLE)
    if ch.get("voice") is True and has_phone:
        out.append(VOICE_ELIGIBLE)
    if out:
        return out
    contactable = ("email", "sms", "voice")
    if all(ch.get(c) is False for c in contactable):
        return [CHANNEL_DNC]
    return [CHANNEL_REVIEW]


def compliance_over_all_candidates(t, candidates: list) -> dict:
    """Most-restrictive-wins across EVERY candidate, not just the chosen one.

    When two distinct master records match one operational row and they
    disagree about a permission, resolving from whichever record the matcher
    happened to rank first can turn a stated denial into permission. The
    identity is ambiguous; the restriction is not. So every candidate votes,
    and any denial anywhere wins.
    """
    out = sr.reconcile_compliance(t, None)
    out["channels"] = {c: None for c in sr.CHANNELS}
    out["findings"] = []
    out["candidate_disagreement"] = []
    for ch_name in sr.CHANNELS:
        states = [getattr(c, f"allow_{ch_name}", None) for c in candidates]
        stated = {s for s in states if s is not None}
        if len(stated) > 1:
            out["candidate_disagreement"].append(ch_name)
        cur = getattr(t, f"allow_{ch_name}", None)
        if cur is False or False in states:
            out["channels"][ch_name] = False
        elif cur is True or True in states:
            out["channels"][ch_name] = True
        if out["channels"][ch_name] is False and cur is not False:
            out["findings"].append({
                "channel": ch_name,
                "reason_code": ("historical_restriction_discovered"
                                if len(stated) <= 1 else
                                "restriction_on_one_of_several_candidates"),
                "current": cur, "historical": False, "resolved": False,
            })
    out["suppressed"] = bool(t.suppressed or any(
        getattr(c, "suppressed", False) for c in candidates))
    out["discovered_restriction"] = bool(out["findings"])
    return out


def rank_key(row: dict):
    """Evidence order, not a manufactured score. Ties are left as ties."""
    disp = DISPOSITION_RANK.get((row["hist_status_reason"] or
                                 row["cur_status"] or "").strip().lower(),
                                1 if (row["hist_status_reason"] or "").strip() else 0)
    action = 1 if (row["hist_last_action"] or "").strip() else 0
    date = row["_activity_dt"]
    recency = date.timestamp() if isinstance(date, datetime) else 0.0
    reach = int(bool(row["_email"])) + int(bool(row["_phone"]))
    return (-disp, -action, -recency, -reach, row["key"])


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--target", required=True)
    ap.add_argument("--source", required=True)
    ap.add_argument("--opportunities", default="")
    ap.add_argument("--out", default="")
    ap.add_argument("--top", type=int, default=25)
    a = ap.parse_args()

    print("READ-ONLY BUSINESS RECONCILIATION")
    print("target:", os.path.basename(a.target))
    print("source:", os.path.basename(a.source))
    if a.opportunities:
        print("opportunities:", os.path.basename(a.opportunities))

    t_rows = read_table(a.target)
    s_rows = read_table(a.source)
    print(f"\ntarget rows {len(t_rows):,}   source rows {len(s_rows):,}")

    # ---- what the operational side actually holds -------------------------
    # The parked activity column is read for REPORTING only. It is explicitly
    # NOT fed to the engine as a current canonical value, because on the Lead
    # row it is not one.
    parked_activity = 0
    for r in t_rows:
        for k, v in r.items():
            if k.strip().lower() == "last activity date" and v is not None \
                    and str(v).strip() != "":
                parked_activity += 1
    print(f"target rows carrying a PARKED last-activity value: {parked_activity}"
          f" of {len(t_rows)}")

    targets = rows_to_records(t_rows)
    parked = []
    for i, t in enumerate(targets):
        parked.append(t.last_contact_date)
        # Model the CURRENT operational state honestly:
        #   the import parked the activity column ...
        t.last_contact_date = None
        #   ... and never mapped any permission column at all.
        t.allow_email = None
        t.allow_bulk_email = None
        t.allow_sms = None
        t.allow_voice = None
        t.suppressed = False
        if not t.key:
            t.key = f"row-{i + 1:03d}"

    sources = rows_to_records(s_rows)

    # ---- does each permission column actually SAY anything? ---------------
    # A column whose every row in the entire master carries the same value
    # carries no information about any individual. It is a default that was
    # never used, not a decision anybody made, and reading it as consent is
    # manufacturing permission out of a column nobody touched. Such a channel
    # is demoted to UNKNOWN for classification - the stated value is still
    # reported, it just stops being treated as evidence.
    informative: dict[str, bool] = {}
    print("\nPERMISSION COLUMN INFORMATION CONTENT (across the whole master)")
    for ch_name in sr.CHANNELS:
        vals = Counter(getattr(s, f"allow_{ch_name}") for s in sources)
        stated = {k: v for k, v in vals.items() if k is not None}
        informative[ch_name] = len(stated) > 1
        desc = "  ".join(f"{perm_word(k)}={v:,}" for k, v in vals.most_common())
        print(f"  {ch_name:<11} {desc}")
        if not informative[ch_name]:
            print(f"  {'':<11} ^ ZERO VARIANCE - one value for all "
                  f"{len(sources):,} rows. This column states no per-person "
                  f"decision, so it is treated as UNKNOWN, never as consent.")
    for s in sources:
        for ch_name in sr.CHANNELS:
            if not informative[ch_name]:
                setattr(s, f"allow_{ch_name}", None)

    index = sr.SourceIndex(sources)
    print(f"index: source_key={len(index.by_source_key):,} "
          f"email={len(index.by_email):,} phone={len(index.by_phone):,} "
          f"name={len(index.by_name):,}")

    findings = sr.reconcile(targets, index)

    # Re-resolve compliance across EVERY candidate. sr.reconcile() resolves it
    # against the single chosen match, which is right whenever the match is
    # unambiguous and wrong exactly when it is not: a record with two matching
    # master contacts that disagree about a permission would take the chosen
    # one's answer, and if that one says Allow while the other says Do Not
    # Allow, a stated denial silently becomes permission. Identity is the
    # ambiguous part; a restriction is not.
    candidate_lists = []
    for t in targets:
        m = sr.match_record(t, index)
        candidate_lists.append(
            [c for c in ([m["matched"]] + list(m["alternates"])) if c])
    n_overridden = 0
    for f, cands in zip(findings, candidate_lists):
        if not cands:
            continue
        before = dict(f["compliance"]["channels"])
        f["compliance"] = compliance_over_all_candidates(f["target"], cands)
        if before != f["compliance"]["channels"]:
            n_overridden += 1
            f["viability_reasons"] = list(f["viability_reasons"]) + [
                "permission_taken_from_most_restrictive_candidate"]
    if n_overridden:
        print(f"\nCOMPLIANCE RE-RESOLVED ACROSS ALL CANDIDATES: {n_overridden} "
              f"record(s) had a denial on a candidate the matcher did not pick")

    # ---- opportunities: corroborating evidence only -----------------------
    # `Contact` in this export is a display NAME, not the contact GUID, so a
    # join on it is name-strength evidence and is treated as such: it can
    # corroborate an outcome the master already states, and it is reported on
    # its own line, but it never silently becomes the authority.
    opp_by_name: dict[str, list[dict]] = defaultdict(list)
    if a.opportunities:
        o_rows = read_table(a.opportunities)
        print(f"opportunity rows {len(o_rows):,}"
              + ("   (NOTE: exactly at the export row cap - may be truncated)"
                 if len(o_rows) in (100000, 99999) else ""))
        for r in o_rows:
            nm = norm_name(str(r.get("Contact") or r.get("Purchaser") or ""))
            if nm:
                opp_by_name[nm].append({
                    "status": str(r.get("Status") or "").strip(),
                    "date": str(r.get("Contract Date") or "").strip(),
                    "total": str(r.get("Contract Total") or "").strip(),
                    "advisor": str(r.get("Sales Advisor") or "").strip(),
                    "type": str(r.get("Contract Type") or "").strip(),
                })

    # ---- per-record business layer ---------------------------------------
    rows = []
    for i, f in enumerate(findings):
        t, s = f["target"], f["source"]
        fills = {x["field"]: x["proposed"] for x in f["enrichment"]["fills"]}
        conflicts = f["enrichment"]["conflicts"]
        ch = f["compliance"]["channels"]

        email = t.email or sr.normalize_email(str(fills.get("email", "") or ""))
        phone = t.phone or str(fills.get("phone", "") or "")
        activity = (s.last_contact_date if s else None) or parked[i]

        name = f"{t.first_name} {t.last_name}".strip()
        opps = opp_by_name.get(norm_name(name), [])
        opp_statuses = [o["status"].lower() for o in opps]
        opp_won = any(x in ("won", "closed won", "complete", "completed")
                      for x in opp_statuses)

        row = {
            "key": t.key or f"row-{i + 1:03d}",
            "current_lead_id": "",          # see REPORT NOTE - not reachable here
            "name": name,
            "master_match_status": f["match_status"],
            "master_match_confidence": f["match_confidence"],
            "master_match_rule": f["match_rule"] or "",
            "historical_contact_guid": (s.source_key if s else ""),
            "alternates": f["alternate_count"],

            # CURRENT (operational, as the Lead row actually holds it)
            "cur_email": t.email,
            "cur_phone": t.phone,
            "cur_address": t.street_address,
            "cur_zip": t.zip_code,
            "cur_status": t.status_reason,
            "cur_last_action": t.last_action,
            "cur_disposition": t.disposition,
            "cur_canonical_last_activity": "",   # parked, never mapped
            "cur_parked_last_activity": ds(parked[i]),
            "cur_perm_email": "UNKNOWN",
            "cur_perm_bulk": "UNKNOWN",
            "cur_perm_sms": "UNKNOWN",
            "cur_perm_voice": "UNKNOWN",

            # HISTORICAL (master)
            "hist_email": (s.email if s else ""),
            # Record.raw is stored with LOWERCASED keys by the adapter.
            "hist_mobile_phone": (str(s.raw.get("mobile phone") or "")
                                  if s else ""),
            "hist_other_phone": (str(s.raw.get("phone") or "") if s else ""),
            "hist_all_phones": ("; ".join(s.norm_phones) if s else ""),
            "hist_perm_raw": ("email=%r bulk=%r sms=%r voice=%r" % (
                s.raw.get("allow emails?"), s.raw.get("do not allow bulk emails"),
                s.raw.get("allow text message?"), s.raw.get("allow phone calls?"))
                if s else ""),
            "hist_address": (s.street_address if s else ""),
            "hist_zip": (s.zip_code if s else ""),
            "hist_last_activity_date": ds(s.last_contact_date if s else None),
            "hist_last_action": (s.last_action if s else ""),
            "hist_status_reason": (s.status_reason if s else ""),
            "hist_sale_made": (s.sale_made if s else ""),
            "hist_last_sold_date": ds(s.last_sold_date if s else None),
            "hist_perm_email": perm_word(s.allow_email if s else None),
            "hist_perm_bulk": perm_word(s.allow_bulk_email if s else None),
            "hist_perm_sms": perm_word(s.allow_sms if s else None),
            "hist_perm_voice": perm_word(s.allow_voice if s else None),

            # RESOLVED (more restrictive wins; UNKNOWN is never consent)
            "resolved_perm_email": perm_word(ch.get("email")),
            "resolved_perm_bulk": perm_word(ch.get("bulk_email")),
            "resolved_perm_sms": perm_word(ch.get("sms")),
            "resolved_perm_voice": perm_word(ch.get("voice")),

            "conflicts": "; ".join(f"{c['field']}: {c['current']!r} vs "
                                   f"{c['historical']!r}" for c in conflicts),
            "conflict_count": len(conflicts),
            "enrichable_fields": "; ".join(sorted(fills)),
            "duplicate_state": ("MULTIPLE_MATCHES" if f["match_status"] ==
                                sr.MULTIPLE_MATCHES else
                                f"alternates={f['alternate_count']}"
                                if f["alternate_count"] else "none"),
            "historical_outcome": "; ".join(
                x for x in [
                    (f"sale_made={s.sale_made}" if s and s.sale_made else ""),
                    (f"last_sold={ds(s.last_sold_date)}"
                     if s and s.last_sold_date else ""),
                    (f"opportunities={len(opps)}" if opps else ""),
                    ("opportunity_won" if opp_won else ""),
                ] if x) or "none_on_record",
            "engine_viability": f["viability"],
            "engine_reasons": "; ".join(f["viability_reasons"]),
            "_email": email,
            "_phone": phone,
            "_activity_dt": activity,
        }
        row["activity_date_used"] = ds(activity)
        row["recommended_channel"] = recommended_channel(
            ch, bool(email), bool(phone))
        wc, wreasons = classify_work(f, ch, bool(email), bool(phone), opp_won)
        row["work_classification"] = wc
        row["work_reasons"] = "; ".join(wreasons)
        # Channels are resolved and reported on their own, so a record held
        # out of the email work list still shows the channels it IS eligible
        # for. This never depends on work_classification.
        row["channel_eligibility"] = "; ".join(channel_eligibility(
            ch, bool(email), bool(phone),
            bool(f["compliance"].get("suppressed")),
            wc == DO_NOT_CONTACT))
        row["candidate_permission_disagreement"] = "; ".join(
            f["compliance"].get("candidate_disagreement", []))
        # Per-channel reason codes. One combined "opted out" sentence hides
        # the case where the channels differ, which is the case that matters.
        row["permission_reasons"] = "; ".join(
            f"{c.upper()} {perm_word(ch.get(c))}" for c in sr.CHANNELS)
        rows.append(row)

    return report(rows, findings, parked_activity, a)


def report(rows: list[dict], findings: list[dict], parked_activity: int,
           a) -> int:
    n = len(rows)
    print("\n" + "=" * 74)
    print("FINAL BUSINESS REPORT")
    print("=" * 74)
    print(f"TOTAL                                 {n}")

    ms = Counter(r["master_match_status"] for r in rows)
    print(f"MASTER HIGH-CONFIDENCE MATCHES        {ms[sr.MATCHED_HIGH_CONFIDENCE]}")
    print(f"REVIEW MATCHES                        {ms[sr.MATCHED_REVIEW]}")
    print(f"NO MATCHES                            {ms[sr.NO_MATCH]}")
    print(f"MULTIPLE MATCHES                      {ms[sr.MULTIPLE_MATCHES]}")

    print()
    for label, col in (("EMAIL", "resolved_perm_email"),
                       ("BULK EMAIL", "resolved_perm_bulk"),
                       ("SMS", "resolved_perm_sms"),
                       ("VOICE", "resolved_perm_voice")):
        c = Counter(r[col] for r in rows)
        print(f"{label + ' ALLOWED':<38}{c['ALLOWED']}")
        print(f"{label + ' DENIED':<38}{c['DENIED']}")
        print(f"{label + ' UNKNOWN':<38}{c['UNKNOWN']}")

    # The permission columns are reported as a COMBINATION as well as per
    # channel, because "82 allowed" on three separate lines invites the reader
    # to assume three independent populations when it may be one.
    print("\nRESOLVED PERMISSION COMBINATIONS (email/bulk/sms/voice)")
    combo = Counter((r["resolved_perm_email"], r["resolved_perm_bulk"],
                     r["resolved_perm_sms"], r["resolved_perm_voice"])
                    for r in rows)
    for k, v in combo.most_common():
        print(f"  {'/'.join(x[0] for x in k)}   "
              f"{' '.join(f'{x:<8}' for x in k)}  {v}")

    # Channel eligibility, reported independently of work classification.
    print("\nCHANNEL ELIGIBILITY (independent of work classification)")
    ce = Counter()
    for r in rows:
        for c_ in r["channel_eligibility"].split("; "):
            if c_:
                ce[c_] += 1
    for k in (EMAIL_ELIGIBLE, SMS_ELIGIBLE, VOICE_ELIGIBLE, CHANNEL_REVIEW,
              CHANNEL_DNC):
        print(f"  {k:<20} {ce[k]}")

    sms_not_email = [r for r in rows
                     if r["resolved_perm_sms"] == "ALLOWED"
                     and r["resolved_perm_email"] == "DENIED"]
    print(f"\nSMS-PERMITTED BUT EMAIL-DENIED   {len(sms_not_email)}")
    for r in sms_not_email:
        print(f"  {r['key']} {r['name']}  {r['permission_reasons']}  "
              f"class={r['work_classification']}  "
              f"channels={r['channel_eligibility']}")
    if not sms_not_email:
        print("  (none in this population - every email-denied record in it "
              "also carries a stated SMS denial)")

    dis = [r for r in rows if r["candidate_permission_disagreement"]]
    print(f"\nCANDIDATE PERMISSION DISAGREEMENTS   {len(dis)}")
    for r in dis:
        print(f"  {r['key']} {r['name']}  channels where two candidate master "
              f"rows disagree: {r['candidate_permission_disagreement']}  "
              f"-> resolved to the restrictive side")

    hist_activity = sum(1 for r in rows if r["hist_last_activity_date"])
    addr = sum(1 for r in rows if r["hist_address"] and not r["cur_address"])
    mob = sum(1 for r in rows if str(r["hist_mobile_phone"]).strip())
    dups = sum(1 for r in rows if r["duplicate_state"] != "none")
    sold = sum(1 for r in rows if r["work_classification"] == ALREADY_RESOLVED)
    opp_only = sum(1 for r in rows if "opportunity_won" in r["historical_outcome"]
                   and r["work_classification"] != ALREADY_RESOLVED)
    print()
    print(f"HISTORICAL ACTIVITY FOUND             {hist_activity}")
    print(f"ADDRESSES RECOVERABLE                 {addr}")
    print(f"MOBILE NUMBERS CONFIRMED              {mob}")
    print(f"DUPLICATES                            {dups}")
    print(f"COMPLETED / SOLD / RESOLVED           {sold}")
    print(f"  of which the master states a sale   "
          f"{sum(1 for r in rows if r['work_classification'] == ALREADY_RESOLVED and (str(r['hist_sale_made']).strip().lower() == 'yes' or r['hist_last_sold_date']))}")
    print(f"NAME-ONLY OPPORTUNITY HITS (reported, {opp_only}")
    print( "  NOT decisive - see note)")

    print()
    wc = Counter(r["work_classification"] for r in rows)
    for k in WORK_CLASSES:
        print(f"{k:<38}{wc[k]}")
    assert sum(wc.values()) == n, "every record must carry exactly one class"

    # ---- conflicts and enrichment ----------------------------------------
    print("\nCONFLICTS (current populated, historical disagrees)")
    cf = Counter()
    for f in findings:
        for c in f["enrichment"]["conflicts"]:
            cf[c["field"]] += 1
    for k, v in cf.most_common() or [("(none)", 0)]:
        print(f"  {k:<28} {v}")

    print("\nBLANK-FILL CANDIDATES (nothing written by this pass)")
    fl = Counter()
    for f in findings:
        for x in f["enrichment"]["fills"]:
            fl[x["field"]] += 1
    for k, v in fl.most_common() or [("(none)", 0)]:
        print(f"  {k:<28} {v}")

    print("\nCOMPLIANCE RESTRICTIONS DISCOVERED IN THE MASTER")
    ch = Counter()
    for f in findings:
        for item in f["compliance"]["findings"]:
            ch[(item["channel"], item["reason_code"])] += 1
    for (c, code), v in ch.most_common() or [(("(none)", ""), 0)]:
        print(f"  {c:<12} {code:<52} {v}")

    # ---- every record that is NOT workable, with the deciding evidence ----
    print("\n" + "=" * 74)
    print("EVERY RECORD HELD BACK, AND THE EVIDENCE THAT HELD IT")
    print("=" * 74)
    held = [r for r in rows if r["work_classification"] not in
            (WORK_NOW_EMAIL, WORK_NOW_VOICE, WORK_LATER)]
    for r in sorted(held, key=lambda x: (x["work_classification"], x["key"])):
        print(f"  {r['key']:<9} {r['name']:<26} {r['work_classification']}")
        print(f"     evidence: {r['work_reasons'] or '-'}")
        print(f"     current status {r['cur_status']!r} action "
              f"{r['cur_last_action']!r}  |  historical status "
              f"{r['hist_status_reason']!r} action {r['hist_last_action']!r}")
        print(f"     stated permissions in master: {r['hist_perm_raw'] or '-'}")
        print(f"     outcome: {r['historical_outcome']}")

    # ---- top N ------------------------------------------------------------
    workable = [r for r in rows
                if r["work_classification"] in (WORK_NOW_EMAIL, WORK_NOW_VOICE,
                                                WORK_LATER)]
    workable.sort(key=rank_key)
    print("\n" + "=" * 74)
    print(f"TOP {a.top} TO WORK FIRST   (of {len(workable)} workable)")
    print("=" * 74)
    for i, r in enumerate(workable[:a.top], 1):
        print(f"{i:>3}. {r['key']:<10} {r['name']:<26} "
              f"{mask_email(r['_email']):<24} {mask_phone(r['_phone'])}")
        print(f"     class {r['work_classification']}  ->  channel "
              f"{r['recommended_channel']}")
        print(f"     match {r['master_match_rule']} "
              f"({r['master_match_confidence']}) guid "
              f"{(r['historical_contact_guid'] or '-')[:8]}")
        print(f"     historical last activity {r['activity_date_used'] or '-'}"
              f"   disposition "
              f"{(r['hist_status_reason'] or r['cur_status'] or '-')!r}"
              f"   action {(r['hist_last_action'] or '-')!r}")
        print(f"     compliance email={r['resolved_perm_email']} "
              f"bulk={r['resolved_perm_bulk']} sms={r['resolved_perm_sms']} "
              f"voice={r['resolved_perm_voice']}")
        print(f"     channels eligible: {r['channel_eligibility']}")
        print(f"     duplicate/conflict {r['duplicate_state']} / "
              f"{r['conflicts'] or 'none'}")
        print(f"     why here: {r['work_reasons'] or '-'}")

    # ---- old qualification vs reconciled ---------------------------------
    # The OLD pass could only see the operational row: no permissions, no
    # canonical activity date. Reconstructed here from the SAME inputs it had,
    # so the delta is explained rather than asserted.
    print("\n" + "=" * 74)
    print("OLD QUALIFICATION (reconstructed from the pre-reconciliation state)")
    print("=" * 74)
    old = {}
    for r in rows:
        if not r["_email"]:
            old[r["key"]] = "EXCLUDED_no_usable_email"
        elif re.search(r"\b(deceased|do not contact|do not call)\b",
                       (r["cur_status"] or "") + " " + (r["cur_email"] or ""),
                       re.I):
            old[r["key"]] = "EXCLUDED_marker_on_operational_row"
        else:
            old[r["key"]] = "READY"
    oc = Counter(old.values())
    for k, v in oc.most_common():
        print(f"  {k:<38} {v}")

    print("\nWHY RECORDS CHANGED")
    moved = Counter()
    for r in rows:
        was = "READY" if old[r["key"]] == "READY" else "EXCLUDED"
        now = r["work_classification"]
        if was == "READY" and now != WORK_NOW_EMAIL:
            moved[(was, now)] += 1
        elif was == "EXCLUDED" and now in (WORK_NOW_EMAIL, WORK_NOW_VOICE,
                                           WORK_LATER):
            moved[(was, now)] += 1
    for (was, now), v in moved.most_common():
        print(f"  {was:<10} -> {now:<32} {v}")

    print("\nREASONS BEHIND EVERY MOVE OFF 'READY'")
    rr = Counter()
    for r in rows:
        if old[r["key"]] == "READY" and r["work_classification"] != WORK_NOW_EMAIL:
            rr[(r["work_classification"], r["work_reasons"] or "-")] += 1
    for (cls, why), v in rr.most_common():
        print(f"  {v:>3}  {cls:<30} {why}")

    # ---- csv --------------------------------------------------------------
    if a.out:
        cols = [k for k in rows[0] if not k.startswith("_")]
        with open(a.out, "w", encoding="utf-8-sig", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            for r in rows:
                w.writerow({k: r[k] for k in cols})
        print(f"\nper-record detail written: {a.out}")

    print("\nREAD-ONLY: no Lead was modified, no message, email, call or "
          "cadence enrolment was produced, and this process held no database "
          "handle.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
